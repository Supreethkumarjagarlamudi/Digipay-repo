import sys
import os
import time
import datetime
import pytest
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Directory setups
current_dir = os.path.dirname(os.path.abspath(__file__))
report_path = os.path.join(os.path.dirname(current_dir), "E2E_Test_Report_Digipay.xlsx")

class PytestResultCollector:
    def __init__(self):
        self.results = []
        self.logs = []
        self.start_time = datetime.datetime.utcnow()
        
    def pytest_runtest_logreport(self, report):
        if report.when == "call":
            duration = round(report.duration, 2)
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Map test names to category names
            test_name = report.nodeid.split("::")[-1]
            category = "Other"
            
            if test_name.startswith("test_role_"):
                category = "Role Selection"
            elif test_name.startswith("test_login_"):
                category = "Login Screen"
            elif test_name.startswith("test_otp_"):
                category = "OTP Verification"
            elif test_name.startswith("test_home_"):
                category = "Customer Home Dashboard"
            elif test_name.startswith("test_discover_"):
                category = "Discover Nearby"
            elif test_name.startswith("test_wallet_"):
                category = "Wallet AI Insights"
            elif test_name.startswith("test_profile_") or test_name.startswith("test_face_") or test_name.startswith("test_notification_"):
                category = "Profile & Subviews"
            elif test_name.startswith("test_merchant_"):
                category = "Merchant Dashboard"
            elif test_name.startswith("test_edit_merchant_"):
                category = "Edit Shop Details"
            elif test_name.startswith("test_statement_"):
                category = "Statements & Date Grouping"
                
            status = report.outcome.upper()
            error_msg = ""
            if report.failed:
                error_msg = str(report.longrepr)
                
            self.results.append({
                "Category": category,
                "Test Name": test_name,
                "Time (sec)": duration,
                "Status": status,
                "Error": error_msg,
                "Timestamp": timestamp
            })
            
            log_msg = f"[{category}] {test_name} \u2192 {status}"
            self.logs.append({
                "Timestamp": timestamp,
                "Level": "INFO" if status == "PASSED" else "ERROR",
                "Message": log_msg
            })

def create_styled_excel(collector):
    end_time = datetime.datetime.utcnow()
    total_duration = round((end_time - collector.start_time).total_seconds(), 2)
    
    results = collector.results
    logs = collector.logs
    
    passed_list = [r for r in results if r["Status"] == "PASSED"]
    failed_list = [r for r in results if r["Status"] == "FAILED"]
    total_tests = len(results)
    passed_count = len(passed_list)
    failed_count = len(failed_list)
    pass_rate = round((passed_count / total_tests * 100), 2) if total_tests > 0 else 0.0
    
    # 1. Summary DataFrame
    summary_data = [{
        "Test Suite": "Digipay iOS App \u2014 Full E2E & Context-Aware Testing",
        "Total Tests": total_tests,
        "Passed": passed_count,
        "Failed": failed_count,
        "Pass Rate %": f"{pass_rate}%",
        "Duration (sec)": total_duration,
        "Start Time": collector.start_time.isoformat() + "Z",
        "End Time": end_time.isoformat() + "Z"
    }]
    df_summary = pd.DataFrame(summary_data)
    
    # 2. Passed Tests DataFrame
    passed_rows = []
    for idx, r in enumerate(passed_list, 1):
        passed_rows.append({
            "No.": idx,
            "Category": r["Category"],
            "Test Name": r["Test Name"],
            "Time (sec)": r["Time (sec)"],
            "Status": "PASSED"
        })
    df_passed = pd.DataFrame(passed_rows)
    if df_passed.empty:
        df_passed = pd.DataFrame(columns=["No.", "Category", "Test Name", "Time (sec)", "Status"])
        
    # 3. Failed Tests DataFrame
    failed_rows = []
    for idx, r in enumerate(failed_list, 1):
        # Keep clean short error summaries
        short_err = r["Error"].split("\n")[-1] if r["Error"] else "Unknown error encountered."
        failed_rows.append({
            "No.": idx,
            "Category": r["Category"],
            "Test Name": r["Test Name"],
            "Error": short_err,
            "Status": "FAILED",
            "Timestamp": r["Timestamp"]
        })
    df_failed = pd.DataFrame(failed_rows)
    if df_failed.empty:
        df_failed = pd.DataFrame(columns=["No.", "Category", "Test Name", "Error", "Status", "Timestamp"])
        
    # 4. Execution Log DataFrame
    df_logs = pd.DataFrame(logs)
    if df_logs.empty:
        df_logs = pd.DataFrame(columns=["Timestamp", "Level", "Message"])
        
    # 5. Test Details DataFrame
    details_rows = []
    for idx, r in enumerate(results, 1):
        err_detail = "None \u2014 test passed successfully." if r["Status"] == "PASSED" else r["Error"]
        details_rows.append({
            "No.": idx,
            "Category": r["Category"],
            "Test Name": r["Test Name"],
            "Status": r["Status"],
            "Error Details": err_detail
        })
    df_details = pd.DataFrame(details_rows)
    if df_details.empty:
        df_details = pd.DataFrame(columns=["No.", "Category", "Test Name", "Status", "Error Details"])
        
    # Write Excel with styling
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_passed.to_excel(writer, sheet_name="Passed Tests", index=False)
        df_failed.to_excel(writer, sheet_name="Failed Tests", index=False)
        df_logs.to_excel(writer, sheet_name="Execution Log", index=False)
        df_details.to_excel(writer, sheet_name="Test Details", index=False)
        
    # Apply styling via openpyxl
    wb = openpyxl.load_workbook(report_path)
    
    # Styles
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cyan_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    soft_red_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    green_text_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # light green
    red_text_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # light red
    
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    normal_font = Font(name="Segoe UI", size=10)
    monospace_font = Font(name="Courier New", size=10)
    
    border_side = Side(style="thin", color="D1D5DB")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    alignment_right = Alignment(horizontal="right", vertical="center")
    
    # Format sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Determine header fill based on sheet
        if sheet_name == "Summary":
            header_fill = navy_fill
        elif sheet_name == "Passed Tests":
            header_fill = cyan_fill
        elif sheet_name == "Failed Tests":
            header_fill = soft_red_fill
        elif sheet_name == "Execution Log":
            header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid") # gray
        else:
            header_fill = navy_fill
            
        # Header formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = alignment_center
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        # Row formatting
        max_row = ws.max_row
        max_col = ws.max_column
        
        for r_idx in range(2, max_row + 1):
            ws.row_dimensions[r_idx].height = 22
            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = alignment_left
                
                # Check status coloring
                if sheet_name in ["Passed Tests", "Failed Tests", "Test Details"]:
                    val_str = str(cell.value)
                    if val_str == "PASSED":
                        cell.fill = green_text_fill
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="15803D")
                        cell.alignment = alignment_center
                    elif val_str == "FAILED":
                        cell.fill = red_text_fill
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
                        cell.alignment = alignment_center
                        
                if sheet_name == "Execution Log":
                    cell.font = monospace_font
                    if c_idx == 2: # Level column
                        cell.alignment = alignment_center
                        if str(cell.value) == "ERROR":
                            cell.font = Font(name="Courier New", size=10, bold=True, color="B91C1C")
                            
                # Align numbers right
                if type(cell.value) in [int, float]:
                    cell.alignment = alignment_right
                    
        # Adjust column widths automatically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(report_path)
    print(f"Styled Excel E2E Test Report successfully saved to: {report_path}")

if __name__ == "__main__":
    print("Initializing Digipay E2E Mobile Appium Test Suite execution...")
    collector = PytestResultCollector()
    
    # Run tests programmatically
    args = [
        "-v",
        os.path.join(current_dir, "test_auth.py"),
        os.path.join(current_dir, "test_customer.py"),
        os.path.join(current_dir, "test_merchant.py")
    ]
    
    exit_code = pytest.main(args, plugins=[collector])
    create_styled_excel(collector)
    
    print("Test execution and Excel report generation completed.")
    sys.exit(exit_code)
