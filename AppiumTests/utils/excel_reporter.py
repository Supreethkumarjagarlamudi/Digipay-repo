"""
excel_reporter.py - Excel Report Generator for Digipay Appium Tests
Produces a rich, multi-sheet Excel workbook with:
  - Summary Dashboard (KPIs, pass/fail pie)
  - Detailed test results table
  - Per-category breakdown sheets
"""

import os
import datetime
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter


# ─── Colour Palette ──────────────────────────────────────────────────────────
BLUE        = "1E40AF"
CYAN        = "0EA5E9"
GREEN       = "16A34A"
RED         = "DC2626"
ORANGE      = "EA580C"
PURPLE      = "7C3AED"
LIGHT_BLUE  = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED   = "FEE2E2"
LIGHT_GRAY  = "F1F5F9"
DARK_GRAY   = "1E293B"
WHITE       = "FFFFFF"
YELLOW      = "F59E0B"

STATUS_FILL = {
    "PASS":  PatternFill("solid", fgColor=LIGHT_GREEN),
    "FAIL":  PatternFill("solid", fgColor=LIGHT_RED),
    "ERROR": PatternFill("solid", fgColor="FFF3CD"),
    "SKIP":  PatternFill("solid", fgColor=LIGHT_GRAY),
}

CATEGORY_COLORS = {
    "Authentication": BLUE,
    "Functional":     CYAN,
    "UI/UX":          PURPLE,
    "Validation":     ORANGE,
    "Unit":           GREEN,
    "Performance":    YELLOW,
    "Security":       RED,
    "Regression":     DARK_GRAY,
    "General":        "6B7280",
}


def _thin_border():
    thin = Side(border_style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _body_font(size=10, bold=False, color=DARK_GRAY):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


class ExcelReporter:
    """Generates a multi-sheet Excel workbook from Pytest Appium results."""

    def __init__(self, results: List[Dict[str, Any]], total_duration: float):
        self.results = results
        self.total_duration = total_duration
        self.wb = openpyxl.Workbook()

    # ── Public API ────────────────────────────────────────────────────────────

    def generate(self, output_path: str):
        self._build_summary_sheet()
        self._build_details_sheet()
        self._build_category_sheets()
        self._build_failures_sheet()
        self.wb.save(output_path)

    # ── Summary Sheet ─────────────────────────────────────────────────────────

    def _build_summary_sheet(self):
        ws = self.wb.active
        ws.title = "📊 Dashboard"
        ws.sheet_view.showGridLines = False

        total    = len(self.results)
        passed   = sum(1 for r in self.results if r["Status"] == "PASS")
        failed   = sum(1 for r in self.results if r["Status"] == "FAIL")
        errors   = sum(1 for r in self.results if r["Status"] == "ERROR")
        skipped  = total - passed - failed - errors
        pass_pct = round((passed / total * 100) if total else 0, 1)

        # ── Banner ────────────────────────────────────────────────────────────
        ws.merge_cells("A1:K2")
        banner = ws["A1"]
        banner.value = "🚀 DIGIPAY iOS — APPIUM E2E TEST REPORT"
        banner.font      = Font(name="Calibri", size=20, bold=True, color=WHITE)
        banner.alignment = _center()
        banner.fill      = PatternFill("solid", fgColor=BLUE)

        ws.merge_cells("A3:K3")
        ts_cell = ws["A3"]
        ts_cell.value     = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Total Duration: {self.total_duration}s"
        ts_cell.font      = _body_font(size=10, color="475569")
        ts_cell.alignment = _center()
        ts_cell.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)

        # ── KPI Cards (row 5-8) ───────────────────────────────────────────────
        kpis = [
            ("Total Tests",    total,    DARK_GRAY, "A5:B8"),
            ("✅ Passed",       passed,   GREEN,     "C5:D8"),
            ("❌ Failed",       failed,   RED,       "E5:F8"),
            ("⚠️ Errors",       errors,   ORANGE,    "G5:H8"),
            ("⏭️ Skipped",      skipped,  "6B7280",  "I5:J8"),
            ("Pass Rate",     f"{pass_pct}%", BLUE, "K5:L8"),
        ]
        for label, value, color, cell_range in kpis:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value     = f"{label}\n{value}"
            cell.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
            cell.alignment = _center()
            cell.fill      = PatternFill("solid", fgColor=color)
            cell.border    = _thin_border()

        # ── Category breakdown table (row 10+) ───────────────────────────────
        self._write_category_table(ws, start_row=10)

        # ── Pie chart ────────────────────────────────────────────────────────
        self._add_pie_chart(ws, data_start=10)

        # ── Column widths ────────────────────────────────────────────────────
        for col_letter, width in zip("ABCDEFGHIJKL", [14]*12):
            ws.column_dimensions[col_letter].width = width
        for row in range(1, 40):
            ws.row_dimensions[row].height = 20

    def _write_category_table(self, ws, start_row: int):
        headers = ["Category", "Total", "Pass", "Fail", "Error", "Pass %"]
        cats: Dict[str, Dict] = {}
        for r in self.results:
            c = r["Category"]
            cats.setdefault(c, {"Total": 0, "PASS": 0, "FAIL": 0, "ERROR": 0})
            cats[c]["Total"] += 1
            cats[c][r["Status"]] += 1

        # Header row
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font      = _header_font(size=10, color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=DARK_GRAY)
            cell.alignment = _center()
            cell.border    = _thin_border()

        for i, (cat, data) in enumerate(cats.items(), start=1):
            row = start_row + i
            pct = round(data["PASS"] / data["Total"] * 100, 1) if data["Total"] else 0
            vals = [cat, data["Total"], data["PASS"], data["FAIL"], data["ERROR"], f"{pct}%"]
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.alignment = _center()
                cell.border    = _thin_border()
                cell.font      = _body_font()
                if col == 4 and data["FAIL"] > 0:
                    cell.fill = STATUS_FILL["FAIL"]
                elif col == 1:
                    color = CATEGORY_COLORS.get(cat, "6B7280")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = _header_font(size=10, color=WHITE)

        # Store last row for chart reference
        self._cat_table_end = start_row + len(cats)
        self._cat_table_start = start_row

    def _add_pie_chart(self, ws, data_start: int):
        chart = PieChart()
        chart.title  = "Test Pass/Fail Distribution"
        chart.style  = 10

        labels  = Reference(ws, min_col=1, min_row=data_start + 1, max_row=self._cat_table_end)
        data    = Reference(ws, min_col=2, min_row=data_start, max_row=self._cat_table_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.shape = 4
        ws.add_chart(chart, "D10")

    # ── Details Sheet ─────────────────────────────────────────────────────────

    def _build_details_sheet(self):
        ws = self.wb.create_sheet("📋 Test Details")
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:J1")
        t = ws["A1"]
        t.value     = "DIGIPAY — Complete Test Results"
        t.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
        t.fill      = PatternFill("solid", fgColor=BLUE)
        t.alignment = _center()

        headers = [
            "Test ID", "Module", "Class", "Test Name", "Category",
            "Status", "Duration (s)", "Error Message", "Timestamp"
        ]
        col_widths = [10, 25, 28, 40, 16, 10, 12, 50, 20]

        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _header_font(color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=DARK_GRAY)
            cell.alignment = _center()
            cell.border    = _thin_border()
            ws.column_dimensions[get_column_letter(col)].width = w

        for row_idx, result in enumerate(self.results, start=3):
            for col_idx, key in enumerate(headers, start=1):
                val  = result.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = _thin_border()
                cell.alignment = _center() if col_idx != 8 else _left()
                cell.font      = _body_font()
                status = result.get("Status", "")
                if col_idx == 6:
                    cell.fill = STATUS_FILL.get(status, PatternFill())
                    cell.font = Font(
                        name="Calibri", size=10, bold=True,
                        color=GREEN if status == "PASS" else RED
                    )
                elif row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

        ws.auto_filter.ref = f"A2:I{len(self.results) + 2}"
        ws.freeze_panes = "A3"

    # ── Per-Category Sheets ───────────────────────────────────────────────────

    def _build_category_sheets(self):
        cats: Dict[str, list] = {}
        for r in self.results:
            cats.setdefault(r["Category"], []).append(r)

        for category, tests in cats.items():
            short = category[:28]
            ws = self.wb.create_sheet(f"🗂 {short}")
            ws.sheet_view.showGridLines = False
            color = CATEGORY_COLORS.get(category, "6B7280")

            ws.merge_cells("A1:G1")
            t = ws["A1"]
            t.value     = f"{category} — Test Results"
            t.font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
            t.fill      = PatternFill("solid", fgColor=color)
            t.alignment = _center()

            headers = ["Test ID", "Test Name", "Status", "Duration (s)", "Error Message", "Timestamp"]
            widths  = [10, 50, 10, 12, 50, 20]
            for col, (h, w) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font      = _header_font(color=WHITE)
                cell.fill      = PatternFill("solid", fgColor=DARK_GRAY)
                cell.alignment = _center()
                cell.border    = _thin_border()
                ws.column_dimensions[get_column_letter(col)].width = w

            for r_idx, t_data in enumerate(tests, start=3):
                row_vals = [
                    t_data["Test ID"], t_data["Test Name"], t_data["Status"],
                    t_data["Duration (s)"], t_data["Error Message"], t_data["Timestamp"]
                ]
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border    = _thin_border()
                    cell.font      = _body_font()
                    cell.alignment = _center() if c_idx != 5 else _left()
                    if c_idx == 3:
                        cell.fill = STATUS_FILL.get(t_data["Status"], PatternFill())
                    elif r_idx % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

            # Mini bar chart
            self._add_mini_bar_chart(ws, tests, category, start_row=r_idx + 2)

    def _add_mini_bar_chart(self, ws, tests, category, start_row):
        passed = sum(1 for t in tests if t["Status"] == "PASS")
        failed = sum(1 for t in tests if t["Status"] == "FAIL")
        ws.cell(row=start_row,     column=1, value="Status")
        ws.cell(row=start_row,     column=2, value="Count")
        ws.cell(row=start_row + 1, column=1, value="Pass")
        ws.cell(row=start_row + 1, column=2, value=passed)
        ws.cell(row=start_row + 2, column=1, value="Fail")
        ws.cell(row=start_row + 2, column=2, value=failed)

        chart = BarChart()
        chart.title   = f"{category} P/F"
        chart.style   = 10
        chart.height  = 8
        chart.width   = 12
        data   = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 2)
        labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, f"D{start_row}")

    # ── Failures Sheet ────────────────────────────────────────────────────────

    def _build_failures_sheet(self):
        failures = [r for r in self.results if r["Status"] in ("FAIL", "ERROR")]
        if not failures:
            return

        ws = self.wb.create_sheet("🔴 Failures")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value     = f"⚠️ FAILED TESTS — {len(failures)} items require attention"
        t.font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
        t.fill      = PatternFill("solid", fgColor=RED)
        t.alignment = _center()

        headers = ["Test ID", "Module", "Test Name", "Category", "Status", "Error Message"]
        widths  = [10, 25, 45, 16, 10, 60]
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _header_font(color=WHITE)
            cell.fill      = PatternFill("solid", fgColor=DARK_GRAY)
            cell.alignment = _center()
            cell.border    = _thin_border()
            ws.column_dimensions[get_column_letter(col)].width = w

        for r_idx, result in enumerate(failures, start=3):
            for c_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=result.get(key, ""))
                cell.border    = _thin_border()
                cell.font      = _body_font()
                cell.alignment = _center() if c_idx != 6 else _left()
                cell.fill      = STATUS_FILL.get(result["Status"], PatternFill())
