#!/usr/bin/env python3
"""
Generate a styled Excel test report from wdio test output.
Usage: python3 scripts/generate_excel_report.py [log_file] [output_excel]
"""

import sys
import re
import json
import os
from datetime import datetime

LOG_FILE = sys.argv[1] if len(sys.argv) > 1 else None
OUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "reports/Digipay_Test_Report.xlsx"
JSON_FILE = "reports/json/test_results.json"

# ── 1. Parse tests from log (preferred) or JSON fallback ─────────────────────

def parse_from_log(log_path):
    """Extract all pass/fail lines from wdio console log."""
    results = []
    suite_map = {
        "TC-SMK": "Smoke Testing Suite",
        "TC-FUN": "Functional Testing Suite",
        "TC-VAL": "Validation Testing Suite",
        "TC-NAV": "Navigation Testing Suite",
        "TC-REG": "Regression Testing Suite",
        "TC-PERF": "Performance Testing Suite",
        "TC-ACC":  "Accessibility Testing Suite",
        "TC-UI":   "UI/UX Testing Suite",
    }
    seen = set()
    with open(log_path, encoding="utf-8") as f:
        for line in f:
            # Match lines like:   ✓ TC-SMK-001 [...]
            m = re.search(r"([✓✗])\s+(TC-\w+-\d+)\s+(\[.*?\]):\s+(.+)", line)
            if not m:
                continue
            symbol, tc_id, meta, title = m.groups()
            if tc_id in seen:
                continue
            seen.add(tc_id)

            # Parse metadata tags
            priority = re.search(r"Priority:\s*(\w+)", meta)
            module   = re.search(r"Module:\s*([\w/]+)", meta)
            feature  = re.search(r"Feature:\s*([^,\]]+)", meta)

            prefix = tc_id.rsplit("-", 1)[0]  # e.g. TC-SMK
            results.append({
                "tc_id":    tc_id,
                "suite":    suite_map.get(prefix, "Unknown Suite"),
                "module":   module.group(1)   if module   else "",
                "priority": priority.group(1) if priority else "",
                "feature":  feature.group(1).strip()  if feature  else "",
                "title":    title.strip(),
                "status":   "PASSED" if symbol == "✓" else "FAILED",
            })
    return results

def parse_from_json(json_path):
    """Fallback: parse the JSON results file produced by the custom reporter."""
    with open(json_path, encoding="utf-8") as f:
        raw = json.load(f)
    results = []
    for item in raw:
        name = item.get("name", "")
        m = re.search(r"(TC-\w+-\d+)\s+(\[.*?\]):\s+(.+)", name)
        if not m:
            continue
        tc_id, meta, title = m.groups()
        priority = re.search(r"Priority:\s*(\w+)", meta)
        module   = re.search(r"Module:\s*([\w/]+)", meta)
        feature  = re.search(r"Feature:\s*([^,\]]+)", meta)
        results.append({
            "tc_id":    tc_id,
            "suite":    item.get("suite", ""),
            "module":   module.group(1)   if module   else "",
            "priority": priority.group(1) if priority else "",
            "feature":  feature.group(1).strip()  if feature  else "",
            "title":    title.strip(),
            "status":   item.get("status", "UNKNOWN"),
        })
    return results

# Decide source
if LOG_FILE and os.path.exists(LOG_FILE):
    tests = parse_from_log(LOG_FILE)
elif os.path.exists(JSON_FILE):
    tests = parse_from_json(JSON_FILE)
else:
    print("ERROR: No log or JSON results file found.")
    sys.exit(1)

if not tests:
    print("ERROR: No test results parsed.")
    sys.exit(1)

print(f"Parsed {len(tests)} unique test results.")

# ── 2. Build Excel with openpyxl ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("openpyxl not installed. Installing...")
    os.system("pip3 install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

wb = openpyxl.Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BG   = "1E1E2E"   # header bg
ACCENT    = "7C3AED"   # purple accent
PASS_BG   = "D1FAE5"   # green tint
FAIL_BG   = "FEE2E2"   # red tint
PASS_FONT = "065F46"
FAIL_FONT = "991B1B"
HEADER_FG = "FFFFFF"
ALT_ROW   = "F5F3FF"   # very light purple for alternating rows
BORDER_C  = "C4B5FD"

thin = Side(style="thin", color=BORDER_C)
thick= Side(style="medium", color=ACCENT)
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def hdr_font(size=11):
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=size)

def body_font(bold=False, color="000000"):
    return Font(name="Calibri", bold=bold, color=color, size=10)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ════════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — SUMMARY DASHBOARD
# ════════════════════════════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "📊 Summary"
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 30
ws_sum.column_dimensions["B"].width = 18
ws_sum.column_dimensions["C"].width = 18
ws_sum.column_dimensions["D"].width = 18
ws_sum.column_dimensions["E"].width = 18
ws_sum.column_dimensions["F"].width = 22

# Title banner
ws_sum.merge_cells("A1:F1")
c = ws_sum["A1"]
c.value = "🧪  DIGIPAY — MOBILE TEST EXECUTION REPORT"
c.font  = Font(name="Calibri", bold=True, color=HEADER_FG, size=18)
c.fill  = fill(DARK_BG)
c.alignment = center()
ws_sum.row_dimensions[1].height = 45

# Subtitle
ws_sum.merge_cells("A2:F2")
c = ws_sum["A2"]
c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Platform: iOS Simulator  |  Framework: Appium + WebdriverIO"
c.font  = Font(name="Calibri", italic=True, color="6D28D9", size=10)
c.fill  = fill("EDE9FE")
c.alignment = center()
ws_sum.row_dimensions[2].height = 22

ws_sum.row_dimensions[3].height = 12  # spacer

# Overall KPIs
total   = len(tests)
passed  = sum(1 for t in tests if t["status"] == "PASSED")
failed  = total - passed
pass_pct= round(passed / total * 100, 1) if total else 0

kpi_headers = ["Total Tests", "Passed ✅", "Failed ❌", "Pass Rate", "Execution Date", "Environment"]
kpi_values  = [total, passed, failed, f"{pass_pct}%",
               datetime.now().strftime("%d %b %Y"), "iOS Simulator (iPhone 17)"]

for col, (h, v) in enumerate(zip(kpi_headers, kpi_values), start=1):
    hc = ws_sum.cell(row=4, column=col, value=h)
    hc.font = hdr_font(10)
    hc.fill = fill(ACCENT)
    hc.alignment = center()
    hc.border = header_border

    vc = ws_sum.cell(row=5, column=col, value=v)
    vc.font = Font(name="Calibri", bold=True, size=14,
                   color=(PASS_FONT if col == 2 else (FAIL_FONT if col == 3 else DARK_BG)))
    vc.fill = fill("FAF5FF")
    vc.alignment = center()
    vc.border = cell_border

ws_sum.row_dimensions[4].height = 22
ws_sum.row_dimensions[5].height = 35

ws_sum.row_dimensions[6].height = 12  # spacer

# Per-suite table
suite_order = [
    "Smoke Testing Suite", "Functional Testing Suite",
    "Validation Testing Suite", "Navigation Testing Suite",
    "Regression Testing Suite", "Performance Testing Suite",
    "Accessibility Testing Suite", "UI/UX Testing Suite",
]

suite_data = {}
for t in tests:
    s = t["suite"]
    if s not in suite_data:
        suite_data[s] = {"pass": 0, "fail": 0}
    suite_data[s]["pass" if t["status"] == "PASSED" else "fail"] += 1

suite_cols = ["Test Suite", "Total", "Passed", "Failed", "Pass Rate", "Status"]
for col, h in enumerate(suite_cols, start=1):
    c = ws_sum.cell(row=7, column=col, value=h)
    c.font = hdr_font(10)
    c.fill = fill("4C1D95")
    c.alignment = center()
    c.border = header_border

for row_i, suite in enumerate(suite_order, start=8):
    data = suite_data.get(suite, {"pass": 0, "fail": 0})
    s_total = data["pass"] + data["fail"]
    s_pass  = data["pass"]
    s_fail  = data["fail"]
    s_rate  = f"{round(s_pass/s_total*100,1)}%" if s_total else "N/A"
    s_status= "✅ PASS" if s_fail == 0 else "❌ FAIL"

    row_fill = fill(ALT_ROW if row_i % 2 == 0 else "FFFFFF")
    for col, val in enumerate([suite, s_total, s_pass, s_fail, s_rate, s_status], start=1):
        c = ws_sum.cell(row=row_i, column=col, value=val)
        c.font = body_font(bold=(col == 1))
        c.fill = row_fill
        c.alignment = center() if col > 1 else left()
        c.border = cell_border
        if col == 3:
            c.font = Font(name="Calibri", bold=True, color=PASS_FONT)
        if col == 4 and val > 0:
            c.font = Font(name="Calibri", bold=True, color=FAIL_FONT)
        if col == 6:
            c.fill = fill(PASS_BG if s_fail == 0 else FAIL_BG)
    ws_sum.row_dimensions[row_i].height = 20

# ════════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — ALL TEST CASES (master view)
# ════════════════════════════════════════════════════════════════════════════════
ws_all = wb.create_sheet("📋 All Test Cases")
ws_all.sheet_view.showGridLines = False

cols_all = ["#", "Test Case ID", "Suite", "Module", "Priority", "Feature", "Test Description", "Status"]
col_widths= [5, 14, 28, 16, 10, 22, 60, 10]

for col, (h, w) in enumerate(zip(cols_all, col_widths), start=1):
    ws_all.column_dimensions[get_column_letter(col)].width = w
    c = ws_all.cell(row=1, column=col, value=h)
    c.font = hdr_font()
    c.fill = fill(DARK_BG)
    c.alignment = center()
    c.border = header_border

ws_all.row_dimensions[1].height = 24
ws_all.freeze_panes = "A2"

for row_i, t in enumerate(tests, start=2):
    is_pass = t["status"] == "PASSED"
    row_bg  = fill(ALT_ROW if row_i % 2 == 0 else "FFFFFF")
    vals = [row_i - 1, t["tc_id"], t["suite"], t["module"],
            t["priority"], t["feature"], t["title"], t["status"]]
    for col, val in enumerate(vals, start=1):
        c = ws_all.cell(row=row_i, column=col, value=val)
        c.font = body_font()
        c.fill = row_bg
        c.alignment = center() if col != 7 else left()
        c.border = cell_border
        if col == 8:  # status column
            c.fill = fill(PASS_BG if is_pass else FAIL_BG)
            c.font = Font(name="Calibri", bold=True,
                          color=PASS_FONT if is_pass else FAIL_FONT)
    ws_all.row_dimensions[row_i].height = 18

# ════════════════════════════════════════════════════════════════════════════════
#  SHEETS 3–10 — Per-suite worksheets
# ════════════════════════════════════════════════════════════════════════════════
suite_sheet_names = {
    "Smoke Testing Suite":         "🔥 Smoke",
    "Functional Testing Suite":    "⚙️ Functional",
    "Validation Testing Suite":    "✅ Validation",
    "Navigation Testing Suite":    "🧭 Navigation",
    "Regression Testing Suite":    "🔁 Regression",
    "Performance Testing Suite":   "⚡ Performance",
    "Accessibility Testing Suite": "♿ Accessibility",
    "UI/UX Testing Suite":         "🎨 UI UX",
}

cols_suite = ["#", "Test Case ID", "Module", "Priority", "Feature", "Test Description", "Status"]
widths_suite=[5, 14, 16, 10, 22, 65, 10]

for suite_name, sheet_label in suite_sheet_names.items():
    suite_tests = [t for t in tests if t["suite"] == suite_name]
    if not suite_tests:
        continue

    ws = wb.create_sheet(sheet_label)
    ws.sheet_view.showGridLines = False

    # Suite header banner
    ws.merge_cells(f"A1:{get_column_letter(len(cols_suite))}1")
    c = ws["A1"]
    c.value = suite_name.upper()
    c.font  = Font(name="Calibri", bold=True, color=HEADER_FG, size=14)
    c.fill  = fill(ACCENT)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(cols_suite, widths_suite), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(10)
        c.fill = fill(DARK_BG)
        c.alignment = center()
        c.border = header_border
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for row_i, t in enumerate(suite_tests, start=3):
        is_pass = t["status"] == "PASSED"
        row_bg  = fill(ALT_ROW if row_i % 2 != 0 else "FFFFFF")
        vals = [row_i - 2, t["tc_id"], t["module"], t["priority"],
                t["feature"], t["title"], t["status"]]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.font = body_font()
            c.fill = row_bg
            c.alignment = center() if col != 6 else left()
            c.border = cell_border
            if col == 7:
                c.fill = fill(PASS_BG if is_pass else FAIL_BG)
                c.font = Font(name="Calibri", bold=True,
                              color=PASS_FONT if is_pass else FAIL_FONT)
        ws.row_dimensions[row_i].height = 18

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_FILE) if os.path.dirname(OUT_FILE) else ".", exist_ok=True)
wb.save(OUT_FILE)
print(f"✅  Excel report saved → {OUT_FILE}")
print(f"    {total} tests  |  {passed} passed  |  {failed} failed  |  {pass_pct}% pass rate")
