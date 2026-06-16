"""
report_generator.py — Digipay E2E Test XLSX Report Generator.

Usage:
    python report_generator.py [--input results.json] [--output-dir .]

Reads pytest-json-report output and produces an Excel workbook with:
  Sheet 1 - Summary         : Run metadata, totals, pass rate
  Sheet 2 - Failed Tests    : Only failed tests with error details
  Sheet 3 - Execution Log   : All tests (INFO/ERROR) with timestamps
  Sheet 4 - Test Details    : Full per-test breakdown

Output filename: E2E_Test_Report_Digipay_<timestamp>.xlsx
"""

import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────
# Color palette (matches PancreaScan report aesthetic)
# ─────────────────────────────────────────────────────────────────
COLOR_HEADER_BG    = "1E3A5F"   # dark navy — header background
COLOR_HEADER_FG    = "FFFFFF"   # white text
COLOR_PASS_BG      = "D6F5E3"   # light green
COLOR_PASS_FG      = "1A7A46"   # dark green
COLOR_FAIL_BG      = "FDE8E8"   # light red
COLOR_FAIL_FG      = "C0392B"   # dark red
COLOR_ALT_ROW      = "F2F7FC"   # alternating row blue-tint
COLOR_SUMMARY_TITLE = "0D2B4E"  # deep navy for summary title
COLOR_BORDER        = "BFCFE0"


def make_border():
    thin = Side(style="thin", color=COLOR_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_font():
    return Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=10)


def header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def cell_font(bold=False, color="000000"):
    return Font(bold=bold, color=color, name="Calibri", size=10)


def pass_fill():
    return PatternFill("solid", fgColor=COLOR_PASS_BG)


def fail_fill():
    return PatternFill("solid", fgColor=COLOR_FAIL_BG)


def alt_row_fill():
    return PatternFill("solid", fgColor=COLOR_ALT_ROW)


# ─────────────────────────────────────────────────────────────────
# Category extraction
# ─────────────────────────────────────────────────────────────────
def extract_category(test_node_id: str) -> str:
    """Derive a human-readable category from the test node path."""
    lower = test_node_id.lower()
    if "landing" in lower:
        return "Landing Page"
    if "login" in lower:
        return "Login Portal"
    if "dashboard" in lower:
        return "Dashboard Portal"
    if "api_unit" in lower or "api" in lower:
        if "health" in lower or "root" in lower:
            return "API Health"
        if "auth" in lower:
            return "Authentication API"
        if "wallet" in lower:
            return "Wallet API"
        if "admin" in lower:
            return "Admin API"
        if "merchant" in lower:
            return "Merchant API"
        return "API Unit Tests"
    if "security" in lower:
        if "auth" in lower:
            return "Auth Security"
        if "jwt" in lower or "tamper" in lower:
            return "JWT Security"
        if "injection" in lower or "input" in lower or "xss" in lower:
            return "Input Validation"
        if "idor" in lower:
            return "IDOR Security"
        if "cors" in lower:
            return "CORS Security"
        if "rate" in lower or "dos" in lower:
            return "Rate Limit / DoS"
        return "Security"
    return "General"


def extract_test_name(node_id: str) -> str:
    """Get just the test function name from the node ID."""
    parts = node_id.split("::")
    return parts[-1] if parts else node_id


def format_error(test: dict) -> str:
    """Extract a clean error message from a test result dict."""
    call = test.get("call", {})
    longrepr = call.get("longrepr", "")
    if longrepr:
        # Trim to first 500 chars for readability
        return str(longrepr)[:500]
    return "No error details available."


# ─────────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────────

def apply_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws, headers: list, row: int = 1):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()


def write_data_row(ws, values: list, row: int, status: str = None, alternate: bool = False):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = make_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        if status == "PASSED":
            if col_idx == len(values):  # Status column
                cell.font = Font(bold=True, color=COLOR_PASS_FG, name="Calibri", size=10)
                cell.fill = pass_fill()
            else:
                cell.font = cell_font()
                cell.fill = pass_fill() if alternate else PatternFill()
        elif status == "FAILED":
            if col_idx == len(values) - 1 or col_idx == len(values):
                cell.font = Font(bold=True, color=COLOR_FAIL_FG, name="Calibri", size=10)
                cell.fill = fail_fill()
            else:
                cell.font = cell_font()
                cell.fill = fail_fill()
        else:
            cell.font = cell_font()
            if alternate:
                cell.fill = alt_row_fill()


# ─────────────────────────────────────────────────────────────────
# Sheet 1: Summary
# ─────────────────────────────────────────────────────────────────
def write_summary_sheet(ws, stats: dict, run_time: float, timestamp: str):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "🧾  DIGIPAY — E2E Test Report Summary"
    title_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_TITLE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Blank row
    ws.append(["", ""])

    rows = [
        ("Report Generated", timestamp),
        ("Total Test Cases", stats["total"]),
        ("✅  Passed", stats["passed"]),
        ("❌  Failed", stats["failed"]),
        ("⚠️  Errors / Skipped", stats.get("error", 0) + stats.get("skipped", 0)),
        ("📊  Pass Rate", f"{stats['pass_rate']:.1f}%"),
        ("⏱  Total Run Time", f"{run_time:.2f}s"),
        ("Frontend URL", "https://harishbalaji826-ops.github.io/digipay-web"),
        ("Backend URL",  "https://web-production-86613.up.railway.app"),
        ("Test Framework", "Selenium 4 + Pytest + pytest-json-report"),
        ("Report Format", "XLSX — 4 Sheets (Summary, Failed, Log, Details)"),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(
            bold=True, name="Calibri", size=10, color="1E3A5F"
        )
        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = Font(name="Calibri", size=10)
        val_cell.alignment = Alignment(horizontal="left")

        if label == "✅  Passed":
            val_cell.font = Font(bold=True, color=COLOR_PASS_FG, name="Calibri", size=10)
        elif label == "❌  Failed":
            val_cell.font = Font(bold=True, color=COLOR_FAIL_FG, name="Calibri", size=10)
        elif label == "📊  Pass Rate":
            rate = stats["pass_rate"]
            color = COLOR_PASS_FG if rate >= 80 else (COLOR_FAIL_FG if rate < 50 else "E67E22")
            val_cell.font = Font(bold=True, color=color, name="Calibri", size=11)

        for col in [1, 2]:
            ws.cell(row=i, column=col).border = make_border()
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = alt_row_fill()


# ─────────────────────────────────────────────────────────────────
# Sheet 2: Failed Tests
# ─────────────────────────────────────────────────────────────────
def write_failed_sheet(ws, failed_tests: list):
    headers = ["No.", "Category", "Test Name", "Error", "Status", "Timestamp"]
    write_header_row(ws, headers)
    apply_column_widths(ws, {
        "A": 6, "B": 22, "C": 38, "D": 75, "E": 10, "F": 22
    })
    ws.row_dimensions[1].height = 20

    if not failed_tests:
        ws.append(["", "", "✅ All tests passed — no failures.", "", "", ""])
        return

    for idx, t in enumerate(failed_tests, start=1):
        row_num = idx + 1
        values = [
            idx,
            t["category"],
            t["test_name"],
            t["error"],
            "FAILED",
            t["timestamp"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fail_fill()
            cell.border = make_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 5:  # Status
                cell.font = Font(bold=True, color=COLOR_FAIL_FG, name="Calibri", size=10)
            else:
                cell.font = Font(color=COLOR_FAIL_FG, name="Calibri", size=10)
        ws.row_dimensions[row_num].height = 60


# ─────────────────────────────────────────────────────────────────
# Sheet 3: Execution Log
# ─────────────────────────────────────────────────────────────────
def write_execution_log_sheet(ws, all_tests: list):
    headers = ["Timestamp", "Level", "Message"]
    write_header_row(ws, headers)
    apply_column_widths(ws, {"A": 22, "B": 8, "C": 100})
    ws.row_dimensions[1].height = 20

    for idx, t in enumerate(all_tests, start=2):
        status = t["status"]
        level = "INFO" if status == "PASSED" else "ERROR"
        duration = t.get("duration", 0)
        msg_parts = [
            f"[{t['category']}] {t['test_name']} → {status}",
            f"in {duration:.2f}s" if duration else "",
        ]
        if status != "PASSED":
            msg_parts.append(f"\n{t['error'][:300]}")
        message = " ".join(p for p in msg_parts if p)

        values = [t["timestamp"], level, message]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if level == "ERROR":
                cell.font = Font(color=COLOR_FAIL_FG, name="Calibri", size=9)
                cell.fill = PatternFill("solid", fgColor="FFF0F0")
            else:
                cell.font = Font(name="Calibri", size=9)
                if idx % 2 == 0:
                    cell.fill = alt_row_fill()
        ws.row_dimensions[idx].height = 30 if status == "PASSED" else 55


# ─────────────────────────────────────────────────────────────────
# Sheet 4: Test Details
# ─────────────────────────────────────────────────────────────────
def write_test_details_sheet(ws, all_tests: list):
    headers = ["No.", "Category", "Test Name", "Status", "Duration (s)", "Error Details"]
    write_header_row(ws, headers)
    apply_column_widths(ws, {
        "A": 6, "B": 22, "C": 42, "D": 10, "E": 14, "F": 75
    })
    ws.row_dimensions[1].height = 20

    for idx, t in enumerate(all_tests, start=2):
        status = t["status"]
        error_detail = t["error"] if status != "PASSED" else "None — test passed successfully."
        duration = t.get("duration", 0)

        values = [
            idx - 1,
            t["category"],
            t["test_name"],
            status,
            f"{duration:.2f}" if duration else "—",
            error_detail,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if status == "PASSED":
                cell.font = Font(
                    bold=(col_idx == 4),
                    color=COLOR_PASS_FG if col_idx == 4 else "333333",
                    name="Calibri", size=10
                )
                cell.fill = pass_fill() if idx % 2 == 0 else PatternFill()
            else:
                cell.font = Font(
                    bold=(col_idx == 4),
                    color=COLOR_FAIL_FG if col_idx in (4, 6) else "333333",
                    name="Calibri", size=10
                )
                cell.fill = fail_fill()

        ws.row_dimensions[idx].height = 50 if status != "PASSED" else 18


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────
def generate_report(input_path: str, output_dir: str) -> str:
    with open(input_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    timestamp_str = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    run_timestamp  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Parse tests ───────────────────────────────────────────────
    tests_raw = report.get("tests", [])
    all_tests = []
    for t in tests_raw:
        node_id   = t.get("nodeid", "")
        status    = t.get("outcome", "unknown").upper()
        duration  = t.get("duration", 0)
        category  = extract_category(node_id)
        test_name = extract_test_name(node_id)
        error     = format_error(t) if status != "PASSED" else ""

        all_tests.append({
            "node_id":   node_id,
            "category":  category,
            "test_name": test_name,
            "status":    status,
            "duration":  duration,
            "error":     error,
            "timestamp": run_timestamp,
        })

    failed_tests = [t for t in all_tests if t["status"] != "PASSED"]

    # ── Compute stats ─────────────────────────────────────────────
    total   = len(all_tests)
    passed  = sum(1 for t in all_tests if t["status"] == "PASSED")
    failed  = sum(1 for t in all_tests if t["status"] == "FAILED")
    errored = sum(1 for t in all_tests if t["status"] == "ERROR")
    skipped = sum(1 for t in all_tests if t["status"] == "SKIPPED")
    pass_rate = (passed / total * 100) if total > 0 else 0.0
    run_time  = report.get("duration", 0)

    stats = {
        "total": total,
        "passed": passed,
        "failed": failed,
        "error": errored,
        "skipped": skipped,
        "pass_rate": pass_rate,
    }

    # ── Build workbook ────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1 — Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary_sheet(ws_summary, stats, run_time, run_timestamp)

    # Sheet 2 — Failed Tests
    ws_failed = wb.create_sheet("Failed Tests")
    write_failed_sheet(ws_failed, failed_tests)

    # Sheet 3 — Execution Log
    ws_log = wb.create_sheet("Execution Log")
    write_execution_log_sheet(ws_log, all_tests)

    # Sheet 4 — Test Details
    ws_details = wb.create_sheet("Test Details")
    write_test_details_sheet(ws_details, all_tests)

    # ── Save ──────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    filename = f"E2E_Test_Report_Digipay_{timestamp_str}.xlsx"
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)

    # ── Print summary to stdout (shown in GitHub Actions log) ─────
    print("\n" + "=" * 60)
    print("   DIGIPAY — E2E Test Report Summary")
    print("=" * 60)
    print(f"  Timestamp   : {run_timestamp}")
    print(f"  Total Tests : {total}")
    print(f"  ✅ Passed   : {passed}")
    print(f"  ❌ Failed   : {failed}")
    if errored or skipped:
        print(f"  ⚠  Other    : {errored + skipped} (errors/skipped)")
    print(f"  📊 Pass Rate: {pass_rate:.1f}%")
    print(f"  ⏱ Run Time : {run_time:.2f}s")
    print("=" * 60)
    print(f"  Report saved: {output_path}")
    print("=" * 60)

    if failed_tests:
        print(f"\n  ❌ Failed Tests ({len(failed_tests)}):")
        for t in failed_tests:
            print(f"     • [{t['category']}] {t['test_name']}")

    print()
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate Digipay E2E XLSX report from pytest-json-report output."
    )
    parser.add_argument(
        "--input", default="results.json",
        help="Path to pytest-json-report JSON file (default: results.json)"
    )
    parser.add_argument(
        "--output-dir", default=".",
        help="Directory to save the XLSX report (default: current directory)"
    )
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    generate_report(args.input, args.output_dir)
